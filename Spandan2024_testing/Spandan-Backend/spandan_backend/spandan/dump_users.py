import django_setup
from django.contrib.auth import get_user_model
import pandas as pd
from openpyxl import load_workbook


book = load_workbook(filename="Students.xlsx")
sh1 = book.active

db = get_user_model()

def generate_random_password(N):
    import random 
    import string
    return ''.join(random.choices(string.ascii_lowercase +
                             string.digits, k=N))

def Sniff_xl(sh):
    dic = {}
    for i in range(2, 2000):
        print("Row ", i, " data :")
        Roll_num = sh.cell(row = i,column =1).value
        print(Roll_num)
 
        email = sh.cell(row = i,column = 2).value
        print(email)
        
        name = sh.cell(row = i,column = 3).value
        print(name)
        
        if(email not in dic.keys()):
            dic[email] = True
            import random
            try:
                u = db.objects.create_user(email,name,name,generate_random_password(random.randint(8,15)),Roll_num,is_active=True)
                print("Created Sucessfully")
            except Exception as e:
                print(e)
                
        else:
            print("Key present")
            
        #print(Roll_num,email,fname,lastname)

Sniff_xl(sh1)



#creates user

#u = db.objects.create_user("mail@mail.com","name","n","passbro")